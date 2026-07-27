"""Read/write project report data (openpyxl — data layer only)."""

from __future__ import annotations

import shutil
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path

import openpyxl

from paths import ORIGINAL_TEMPLATE, WORKING_DATA, assert_original_readonly, ensure_dirs

ATTENDANCE_WORK = "เข้าปฎิบัติงาน"
NON_WORK_ATTENDANCE = frozenset(
    {
        "วันเสาร์",
        "วันอาทิตย์",
        "วันหยุดนักขัตฤกษ์",
        "วันเข้าปฎิบัตงาน",
        "วันไม่เข้าปฎิบัตงาน",
    }
)
COL_DATE = 1
COL_ATTENDANCE = 2
COL_JOB_CODE = 3
COL_WORK_TYPE = 4
COL_DETAIL = 5
HEADER_ROW = 7


@dataclass
class ReportEntry:
    event_date: date
    attendance: str
    job_code: str
    work_type: str
    detail: str

    @property
    def remark(self) -> str:
        return self.detail.strip()

    @property
    def is_work_day(self) -> bool:
        if self.attendance in NON_WORK_ATTENDANCE:
            return False
        return self.attendance == ATTENDANCE_WORK and bool(self.detail.strip())


def _guard_writable(path: Path) -> Path:
    resolved = path.resolve()
    if resolved == ORIGINAL_TEMPLATE.resolve():
        raise PermissionError(
            "ห้ามเขียนทับไฟล์ต้นฉบับ — ใช้ 02-working/report-data.xlsx แทน"
        )
    return resolved


def init_working_data() -> Path:
    """Create report-data.xlsx from template copy if missing."""
    assert_original_readonly()
    ensure_dirs()
    if not WORKING_DATA.exists():
        shutil.copy2(ORIGINAL_TEMPLATE, WORKING_DATA)
        print(f"Created data file: {WORKING_DATA}")
    return WORKING_DATA


def _coerce_date(value: object) -> date | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def month_sheet_name(target: date) -> str:
    return target.strftime("%B")


def row_for_day(day: int) -> int:
    return HEADER_ROW + day


def find_row_for_date(ws, target: date, *, path: Path | None = None) -> int | None:
    first_row = HEADER_ROW + 1

    if path is not None:
        data_ws = openpyxl.load_workbook(path, data_only=True)[ws.title]
        anchor = _coerce_date(data_ws.cell(first_row, COL_DATE).value)
        if (
            anchor
            and anchor.year == target.year
            and anchor.month == target.month
            and target.day >= anchor.day
        ):
            return first_row + (target.day - anchor.day)

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        cell_date = _coerce_date(ws.cell(row, COL_DATE).value)
        if cell_date == target:
            return row
    return None


def _resolve_sheet_dates(ws) -> dict[int, date]:
    dates: dict[int, date] = {}
    anchor_row: int | None = None
    anchor_date: date | None = None

    for row in range(HEADER_ROW + 1, ws.max_row + 1):
        raw = ws.cell(row, COL_DATE).value
        if isinstance(raw, str) and raw.startswith("="):
            if anchor_row is not None and anchor_date is not None:
                dates[row] = date.fromordinal(
                    anchor_date.toordinal() + (row - anchor_row)
                )
            continue
        cell_date = _coerce_date(raw)
        if cell_date is None:
            continue
        dates[row] = cell_date
        if anchor_row is None:
            anchor_row = row
            anchor_date = cell_date

    if anchor_row is not None and anchor_date is not None:
        anchor_month = anchor_date.month
        anchor_year = anchor_date.year
        for row in range(anchor_row, ws.max_row + 1):
            if row not in dates:
                candidate = date.fromordinal(
                    anchor_date.toordinal() + (row - anchor_row)
                )
                if candidate.month != anchor_month or candidate.year != anchor_year:
                    break
                dates[row] = candidate
    return dates


def read_month_entries(path: Path, sheet_name: str) -> list[ReportEntry]:
    path = Path(path)
    wb = openpyxl.load_workbook(path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]

    ws_formulas = openpyxl.load_workbook(path, data_only=False)[sheet_name]
    has_formula_dates = any(
        isinstance(ws_formulas.cell(row, COL_DATE).value, str)
        and str(ws_formulas.cell(row, COL_DATE).value).startswith("=")
        for row in range(HEADER_ROW + 1, min(HEADER_ROW + 5, ws_formulas.max_row + 1))
    )
    if has_formula_dates:
        row_dates = _resolve_sheet_dates(ws_formulas)
    else:
        row_dates = {
            row: d
            for row in range(HEADER_ROW + 1, ws.max_row + 1)
            if (d := _coerce_date(ws.cell(row, COL_DATE).value)) is not None
        }

    entries: list[ReportEntry] = []
    for row, event_date in sorted(row_dates.items()):
        attendance = ws.cell(row, COL_ATTENDANCE).value
        attendance = attendance.strip() if isinstance(attendance, str) else ""
        job_code = ws.cell(row, COL_JOB_CODE).value
        job_code = job_code.strip() if isinstance(job_code, str) else ""
        work_type = ws.cell(row, COL_WORK_TYPE).value
        work_type = work_type.strip() if isinstance(work_type, str) else ""
        detail = ws.cell(row, COL_DETAIL).value
        detail = detail.strip() if isinstance(detail, str) else ""
        entries.append(
            ReportEntry(
                event_date=event_date,
                attendance=attendance,
                job_code=job_code,
                work_type=work_type,
                detail=detail,
            )
        )
    return entries


def read_work_entries(
    path: Path | None = None,
    *,
    sheet_name: str | None = None,
    from_date: date | None = None,
    to_date: date | None = None,
) -> list[ReportEntry]:
    path = Path(path or WORKING_DATA)
    if sheet_name:
        sheets = [sheet_name]
    elif from_date and to_date and from_date.month == to_date.month:
        sheets = [month_sheet_name(from_date)]
    else:
        raise ValueError("Provide sheet_name or a single-month from_date/to_date range")

    entries = read_month_entries(path, sheets[0])
    if from_date:
        entries = [e for e in entries if e.event_date >= from_date]
    if to_date:
        entries = [e for e in entries if e.event_date <= to_date]
    return [e for e in entries if e.is_work_day]


def write_entry(ws, row: int, entry: ReportEntry) -> None:
    ws.cell(row, COL_ATTENDANCE, entry.attendance)
    ws.cell(row, COL_JOB_CODE, entry.job_code)
    ws.cell(row, COL_WORK_TYPE, entry.work_type)
    ws.cell(row, COL_DETAIL, entry.detail)


def write_entries(
    path: Path | None,
    sheet_name: str,
    entries: list[ReportEntry],
) -> int:
    path = _guard_writable(Path(path or WORKING_DATA))
    wb = openpyxl.load_workbook(path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet not found: {sheet_name}")
    ws = wb[sheet_name]
    written = 0
    for entry in entries:
        row = find_row_for_date(ws, entry.event_date, path=path)
        if row is None:
            raise ValueError(f"No row for date {entry.event_date} in sheet {sheet_name}")
        write_entry(ws, row, entry)
        written += 1
    wb.save(path)
    return written
